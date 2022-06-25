#PRINTING LINE FORMAT
from openpyxl import Workbook
from openpyxl import load_workbook
import os
def WorkBookCreation(decade):
    workbook = Workbook()
    workbook.save(filename=decade+".xlsx")
    return workbook

def SheetCreation(year,workbook,Firstsheet):
    if Firstsheet:
        sheet = workbook.active
    else:
        sheet = workbook.create_sheet(str(year))
    RaceHeader = ["Driver","Average Finish","Delta","Average per-race Delta","Standard Deviation","Median Finish","Finishing rate","PointsPerRace","Scoring Ratio","Scoring Contribution","Position Normalized Scoring Ratio","Position Normalized scoring Contribution","Points","Ranking","Position Normalized Points","Position Nomalized Ranking"]
    sheet.append(RaceHeader)
    return sheet

    